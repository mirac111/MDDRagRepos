"""
SUT XLSX Parser for RAG
========================

SUT (Sağlık Uygulama Tebliği) Excel dosyalarını RAG sistemi için chunk'lara ayırır.

ÖZELLIKLER:
- Character-level strikethrough detection (openpyxl rich text)
- (Mülga, Değişik, Ek) varyant temizleme
- Horizontal merged header support
- EK kodu ve liste adı extraction (ilk 2 satır)
- Row-level meaningful content check

PATTERN:
- sut.py (DOCX) ile aynı output formatı
- table.py'den farklı: tek content_with_weight field

KULLANIM:
    from rag.app.table_sut import chunk
    
    chunks = chunk(
        filename="EK-1C.xlsx",
        binary=file_binary,
        callback=progress_callback
    )

OUTPUT FORMAT:
    [
        {
            "docnm_kwd": "EK-1C.xlsx",
            "title_tks": [...],
            "title_sm_tks": [...],
            "content_with_weight": "EK-1/C istisnai sağlık...; SIRA NO:1; KODU:P123; ...",
            "content_ltks": [...],
            "content_sm_ltks": [...]
        },
        ...
    ]
"""

import copy
import re
import unicodedata
from io import BytesIO
from typing import List, Dict, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from rag.nlp import rag_tokenizer


# ============================================================================
# REGEX PATTERNS & TEXT CLEANING
# ============================================================================

class RegexPatterns:
    """SUT'a özel regex pattern tanımlamaları"""
    
    # Varyant pattern'leri
    VARIANT = re.compile(
        r'\s*\(\s*(Mülga|Değişik|Ek)\s*:[^)]*\)\s*',
        re.IGNORECASE
    )
    
    VARIANT_UNCLOSED = re.compile(
        r'\s*\(\s*(Mülga|Değişik|Ek)\s*:.*$',
        re.IGNORECASE
    )
    
    YURURLUK = re.compile(
        r'\s*\(?\s*Yürürlük\s*:.*$',
        re.IGNORECASE
    )
    
    MD_YURURLUK = re.compile(
        r'\s*\b\d+\s*md\.\s*Yürürlük\s*:.*$',
        re.IGNORECASE
    )


class TextCleaner:
    """Text temizleme işlemleri"""
    
    @staticmethod
    def clean_variants(text: str) -> str:
        """
        (Mülga:...), (Değişik:...), (Ek:...), (Yürürlük:...) gibi 
        varyant bilgilerini temizler.
        
        Özel durum: "27 (Değişik:...)" → "27" kalır
        """
        if not text:
            return text
        
        text = re.sub(r'\s+', ' ', text)
        
        # Başta sayı varsa kaydet
        leading_number = None
        number_match = re.match(r'^(\d+)\s*', text)
        if number_match:
            leading_number = number_match.group(1)
        
        # Varyantları temizle
        text = RegexPatterns.VARIANT.sub(' ', text)
        text = RegexPatterns.VARIANT_UNCLOSED.sub(' ', text)
        text = RegexPatterns.YURURLUK.sub(' ', text)
        text = RegexPatterns.MD_YURURLUK.sub(' ', text)
        
        # Kod listelerini temizle: "(P618690, P621410... kodlu işlemler hariç)"
        text = re.sub(r'\s*\([P\d,.\s]+kodlu işlemler[^)]*\)\s*', ' ', text, flags=re.IGNORECASE)
        
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        
        # Eğer temizlendikten sonra boş kaldıysa ve başta sayı vardıysa, sayıyı geri ver
        if not text and leading_number:
            return leading_number
        
        return text
    
    @staticmethod
    def is_only_variant(text: str) -> bool:
        """
        Text sadece varyant içeriyor mu kontrol eder.
        Örnek: "(Mülga:RG-24/12/2014-29215 / 18-a md. Yürürlük: 01/01/2015)"
        """
        if not text or not text.strip():
            return False
        
        cleaned = TextCleaner.clean_variants(text)
        return not cleaned or not cleaned.strip()


# ============================================================================
# EXCEL CELL STYLE ANALYSIS
# ============================================================================

class CellStyleAnalyzer:
    """Excel cell formatlamasını analiz eder"""
    
    @staticmethod
    def is_strikethrough(cell) -> bool:
        """Cell'in üstü çizili mi kontrol eder"""
        try:
            if hasattr(cell, 'font') and cell.font:
                return bool(cell.font.strike)
            return False
        except:
            return False
    
    @staticmethod
    def is_empty(cell) -> bool:
        """Cell boş mu kontrol eder"""
        if cell is None:
            return True
        if isinstance(cell, MergedCell):
            return True
        if cell.value is None:
            return True
        if str(cell.value).strip() == "":
            return True
        return False


# ============================================================================
# HEADER DETECTION
# ============================================================================

class HeaderDetector:
    """Excel header'larını tespit eder (horizontal merged support ile)"""
    
    @staticmethod
    def detect_headers(sheet) -> tuple:
        """
        Sheet'ten header'ları ve header satır sayısını tespit eder.
        
        Excel yapısı:
        - Row 0: EK kodu
        - Row 1: Liste adı
        - Row 2: Column headers (merged olabilir)
        
        Returns:
            (headers: List[Dict], header_rows: int)
            headers format: [
                {
                    'name': str,
                    'is_real': bool,
                    'index': int,
                    'is_merged': bool,
                    'merged_span': (start, end) or None,
                    'is_first_in_merge': bool
                },
                ...
            ]
        """
        rows = list(sheet.rows)
        if len(rows) < 3:
            return [], 0
        
        header_row = rows[2]
        headers = []
        processed_cols = set()
        header_row_idx = 2
        excel_header_row_num = header_row_idx + 1  # Excel'de 1-based
        
        # Header satırındaki merged range'leri bul
        merged_ranges_in_header = []
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= excel_header_row_num <= merged_range.max_row:
                start_col_idx = merged_range.min_col - 1  # 0-based
                end_col_idx = merged_range.max_col - 1
                merged_ranges_in_header.append((start_col_idx, end_col_idx))
        
        # Her kolonu işle
        for i in range(len(header_row)):
            if i in processed_cols:
                continue
            
            # Bu kolon merged bir range'in içinde mi?
            is_merged = False
            merged_span = None
            
            for start_col, end_col in merged_ranges_in_header:
                if start_col <= i <= end_col:
                    is_merged = True
                    merged_span = (start_col, end_col)
                    break
            
            if is_merged:
                # Merged range - tek header adı tüm kolonlar için
                start_col, end_col = merged_span
                
                # Header adını bul (ilk dolu cell)
                header_name = None
                for check_col in range(start_col, end_col + 1):
                    if check_col < len(header_row):
                        check_cell = header_row[check_col]
                        if not CellStyleAnalyzer.is_empty(check_cell):
                            header_value = str(check_cell.value).strip()
                            if header_value:
                                header_name = header_value
                                break
                
                if not header_name:
                    header_name = f"Column_{start_col + 1}"
                
                # Her kolon için aynı header ekle
                for col_idx in range(start_col, end_col + 1):
                    headers.append({
                        'name': header_name,
                        'is_real': bool(header_name and not header_name.startswith("Column_")),
                        'index': col_idx,
                        'is_merged': True,
                        'merged_span': (start_col, end_col),
                        'is_first_in_merge': (col_idx == start_col)
                    })
                    processed_cols.add(col_idx)
            else:
                # Normal cell
                cell = header_row[i]
                if not CellStyleAnalyzer.is_empty(cell):
                    header_value = str(cell.value).strip()
                    if header_value:
                        headers.append({
                            'name': header_value,
                            'is_real': True,
                            'index': i,
                            'is_merged': False,
                            'merged_span': None,
                            'is_first_in_merge': False
                        })
                    else:
                        headers.append({
                            'name': f"Column_{i + 1}",
                            'is_real': False,
                            'index': i,
                            'is_merged': False,
                            'merged_span': None,
                            'is_first_in_merge': False
                        })
                else:
                    headers.append({
                        'name': f"Column_{i + 1}",
                        'is_real': False,
                        'index': i,
                        'is_merged': False,
                        'merged_span': None,
                        'is_first_in_merge': False
                    })
                processed_cols.add(i)
        
        return headers, 3  # İlk 3 satır header bilgisi


# ============================================================================
# CELL PROCESSING
# ============================================================================

class CellProcessor:
    """Cell içeriğini işler - character-level strikethrough support"""
    
    @staticmethod
    def extract_segments_from_cell(cell) -> List[Dict]:
        """
        Cell'den segment'leri çıkarır (DOCX run mantığı gibi).
        
        Returns:
            List[{'text': str, 'isStrikethrough': bool}]
        """
        segments = []
        
        if CellStyleAnalyzer.is_empty(cell):
            return segments
        
        try:
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            
            # Rich text mi kontrol et
            if isinstance(cell.value, CellRichText):
                # Rich text - her text block bir segment
                for text_block in cell.value:
                    if isinstance(text_block, TextBlock):
                        text = text_block.text if hasattr(text_block, 'text') else str(text_block)
                        
                        # TextBlock'un strike özelliği
                        is_strike = False
                        if hasattr(text_block, 'font') and text_block.font:
                            is_strike = bool(text_block.font.strike)
                        
                        if text.strip():
                            segments.append({
                                'text': text,
                                'isStrikethrough': is_strike
                            })
                    
                    elif isinstance(text_block, str):
                        # Plain string - strikethrough değil
                        if text_block.strip():
                            segments.append({
                                'text': text_block,
                                'isStrikethrough': False
                            })
            else:
                # Normal cell (rich text değil)
                text = str(cell.value).strip()
                if text:
                    # Cell-level strikethrough
                    is_strike = CellStyleAnalyzer.is_strikethrough(cell)
                    segments.append({
                        'text': text,
                        'isStrikethrough': is_strike
                    })
        
        except Exception:
            # Rich text parse hatası - normal cell olarak işle
            text = str(cell.value).strip() if cell.value else ""
            if text:
                is_strike = CellStyleAnalyzer.is_strikethrough(cell)
                segments.append({
                    'text': text,
                    'isStrikethrough': is_strike
                })
        
        return segments
    
    @staticmethod
    def process_cell(cell, sheet=None) -> Optional[str]:
        """
        Cell'i işler:
        1. MergedCell ise gerçek cell'i bul
        2. Segment'leri çıkar
        3. Strikethrough OLMAYAN segment'leri filtrele
        4. Varyantları temizle
        5. Birleştir
        
        Returns:
            Temizlenmiş text veya None
        """
        if cell is None:
            return None
        
        actual_cell = cell
        
        # MergedCell ise gerçek cell'i bul
        if isinstance(cell, MergedCell):
            if sheet is None:
                return None
            
            try:
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        start_row = merged_range.min_row
                        start_col = merged_range.min_col
                        actual_cell = sheet.cell(row=start_row, column=start_col)
                        break
            except:
                return None
        
        # Boş kontrolü
        if CellStyleAnalyzer.is_empty(actual_cell):
            return None
        
        # Cell-level strikethrough kontrolü (tüm cell strikethrough ise atla)
        if CellStyleAnalyzer.is_strikethrough(actual_cell):
            return None
        
        # Raw text al
        raw_text = str(actual_cell.value).strip() if actual_cell.value else ""

        #if raw_text:
            #import unicodedata
            # Önce küçük harfe çevir (Türkçe-aware)
            #raw_text = raw_text.lower()
            # Sonra NFC normalize et (combining characters'ı düzelt)
            #raw_text = unicodedata.normalize('NFC', raw_text)
        
        if not raw_text:
            return None
        
        # Sadece varyant mı kontrol et
        if TextCleaner.is_only_variant(raw_text):
            return None
         
        # Varyantları temizle
        cleaned_text = TextCleaner.clean_variants(raw_text)
        
        if not cleaned_text or not cleaned_text.strip():
            return None
        
        return cleaned_text


# ============================================================================
# ROW PROCESSING
# ============================================================================

class RowProcessor:
    """Satır işleme - row-level meaningful content check"""
    
    @staticmethod
    def _row_has_any_meaningful_content(row, headers: List[Dict], sheet, row_idx: int) -> bool:
        """
        Row'da anlamlı içerik var mı kontrol eder.
        En az bir cell'de processed value varsa True döner.
        """
        for col_idx in range(len(headers)):
            if col_idx >= len(row):
                break
            
            cell = row[col_idx]
            processed_value = CellProcessor.process_cell(cell, sheet)
            
            if processed_value and processed_value.strip():
                return True
        
        return False
    
    @staticmethod
    def _get_merged_cell_span(sheet, row_idx: int, col_idx: int) -> Optional[tuple]:
        """
        Merged cell'in span'ini döner (horizontal only)
        Returns: (start_col_idx, end_col_idx) veya None
        """
        try:
            cell_coord = sheet.cell(row=row_idx + 1, column=col_idx + 1).coordinate
            for merged_range in sheet.merged_cells.ranges:
                if cell_coord in merged_range:
                    # Sadece horizontal merged (aynı satırda)
                    if merged_range.min_row == merged_range.max_row:
                        return (merged_range.min_col - 1, merged_range.max_col - 1)
            return None
        except:
            return None
    
    @staticmethod
    def _detect_merged_content(row, headers: List[Dict], sheet, row_idx: int) -> Dict:
        """
        Birleşik kolonları tespit eder ve mapping döner.
        
        Returns:
            {
                'is_merged': bool,
                'merged_cols': [(start_idx, end_idx, content), ...],
                'normal_cols': [(col_idx, content), ...]
            }
        """
        result = {
            'is_merged': False,
            'merged_cols': [],
            'normal_cols': []
        }
        
        processed_cols = set()
        excel_row_num = row_idx + 1
        
        # Bu satırdaki horizontal merged range'leri bul
        merged_ranges_in_row = []
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row == excel_row_num and merged_range.max_row == excel_row_num:
                start_col_idx = merged_range.min_col - 1
                end_col_idx = merged_range.max_col - 1
                merged_ranges_in_row.append((start_col_idx, end_col_idx))
        
        # Her kolonu işle
        for col_idx in range(len(headers)):
            if col_idx in processed_cols:
                continue
                
            if col_idx >= len(row):
                break
            
            # Bu kolon merged bir range'in içinde mi?
            is_in_merged = False
            merged_span = None
            
            for start_col, end_col in merged_ranges_in_row:
                if start_col <= col_idx <= end_col:
                    is_in_merged = True
                    merged_span = (start_col, end_col)
                    break
            
            if is_in_merged:
                # Merged cell bulundu
                start_col, end_col = merged_span
                
                # İçeriği bul (ilk dolu cell)
                content_found = None
                for check_col in range(start_col, end_col + 1):
                    if check_col < len(row):
                        check_cell = row[check_col]
                        if not CellStyleAnalyzer.is_empty(check_cell):
                            processed_value = CellProcessor.process_cell(check_cell, sheet)
                            if processed_value:
                                content_found = processed_value
                                break
                
                if content_found:
                    result['is_merged'] = True
                    result['merged_cols'].append((start_col, end_col, content_found))
                
                # Bu range'deki tüm kolonları işlenmiş işaretle
                for i in range(start_col, end_col + 1):
                    processed_cols.add(i)
            else:
                # Normal cell
                cell = row[col_idx]
                processed_value = CellProcessor.process_cell(cell, sheet)
                result['normal_cols'].append((col_idx, processed_value))
                processed_cols.add(col_idx)
        
        return result
    
    @staticmethod
    def process_row(row, headers: List[Dict], sheet, ek_code: str = "", list_name: str = "", row_idx: int = 0) -> Optional[str]:
        """
        Bir satırı işler ve content string oluşturur.
        
        Args:
            row: Excel satırı
            headers: Column headers (Dict list)
            sheet: Sheet objesi
            ek_code: EK kodu (örn: "EK-1/C")
            list_name: Liste adı (örn: "istisnai sağlık hizmetleri listesi")
            row_idx: Satır index'i (merged cell kontrolü için)
        
        Returns:
            Content string veya None
        """
        # Row'da anlamlı içerik var mı kontrol et
        has_meaningful = RowProcessor._row_has_any_meaningful_content(row, headers, sheet, row_idx)
        
        if not has_meaningful:
            return None
        
        # Merged content tespit et
        content_info = RowProcessor._detect_merged_content(row, headers, sheet, row_idx)
        
        # Merged header groups (header satırındaki merged'ler)
        merged_header_groups = {}
        for col_idx in range(len(headers)):
            header = headers[col_idx]
            if header['is_merged'] and header['is_first_in_merge']:
                start_col, end_col = header['merged_span']
                merged_header_groups[start_col] = {
                    'header_name': header['name'],
                    'span': (start_col, end_col),
                    'contents': []
                }
        
        # Tüm kolonları topla
        all_columns = []  # [(index, content_string), ...]
        has_meaningful_data = False
        processed_header_groups = set()
        
        # Merged kolonları işle (row içindeki merged cell'ler)
        if content_info['is_merged']:
            for start_idx, end_idx, content in content_info['merged_cols']:
                # Merged header'ları topla
                merged_header_names = []
                for i in range(start_idx, end_idx + 1):
                    if i < len(headers) and headers[i]['is_real']:
                        merged_header_names.append(headers[i]['name'])
                
                if merged_header_names:
                    merged_header_str = " ve ".join(merged_header_names)
                    all_columns.append((start_idx, f"{merged_header_str}:{content}"))
                    
                    # SIRA NO değilse meaningful data var
                    if not any(h.upper() == 'SIRA NO' for h in merged_header_names):
                        has_meaningful_data = True
        
        # Normal kolonları işle
        for col_idx, processed_value in content_info['normal_cols']:
            if col_idx >= len(headers):
                continue
            
            header_info = headers[col_idx]
            
            # Sadece gerçek header'ları işle
            if not header_info['is_real']:
                continue
            
            # Header merged mi? (header satırında merged)
            if header_info['is_merged']:
                start_col, end_col = header_info['merged_span']
                
                # Bu group zaten işlendi mi?
                if start_col in processed_header_groups:
                    continue
                
                # Merged header group'u işle
                if start_col in merged_header_groups:
                    group_info = merged_header_groups[start_col]
                    
                    # Bu group'taki tüm değerleri topla
                    for check_col_idx in range(start_col, end_col + 1):
                        for norm_col_idx, norm_content in content_info['normal_cols']:
                            if norm_col_idx == check_col_idx and norm_content:
                                group_info['contents'].append(norm_content)
                    
                    # Birleştir
                    if group_info['contents']:
                        combined_content = " ve ".join(group_info['contents'])
                        all_columns.append((start_col, f"{group_info['header_name']}:{combined_content}"))
                        
                        if group_info['header_name'].upper() != 'SIRA NO':
                            has_meaningful_data = True
                    
                    processed_header_groups.add(start_col)
            else:
                # Normal header
                header_name = header_info['name']
                
                if processed_value:
                    all_columns.append((col_idx, f"{header_name}:{processed_value}"))
                    
                    if header_name.upper() != 'SIRA NO':
                        has_meaningful_data = True
                else:
                    # Null değer
                    #all_columns.append((col_idx, f"{header_name}: null"))
                    all_columns.append((col_idx, f"{header_name}:yok"))
        
        # Anlamlı data yoksa atla
        if not has_meaningful_data:
            return None
        
        # Hiç içerik yoksa atla
        if not all_columns:
            return None
        
        # Kolonları INDEX SIRASINA göre sırala
        all_columns.sort(key=lambda x: x[0])
        
        # Content string'leri al
        row_content = [content for _, content in all_columns]
        
        # Content oluştur: EK kodu + liste adı + detaylar
        content_parts = []
        
        if ek_code and list_name:
            content_parts.append(f"{ek_code} {list_name}")
        elif ek_code:
            content_parts.append(ek_code)
        elif list_name:
            content_parts.append(list_name)
        
        # Row içeriğini ekle
        content_parts.append('; '.join(row_content))
        
        # Final content string
        return '; '.join(content_parts)


# ============================================================================
# MAIN SUT XLSX PARSER
# ============================================================================

class SutXlsx:
    """
    SUT XLSX parser - RAG integration
    Pattern: sut.py (DOCX) benzeri
    """
    
    def __init__(self):
        pass
    
    def __call__(self, filename: str, binary: bytes = None, from_page: int = 0, to_page: int = 100000) -> List[str]:
        """
        XLSX dosyasını parse eder ve content string listesi döner.
        
        Args:
            filename: Dosya adı (binary varsa göz ardı edilir)
            binary: Dosya binary (BytesIO için)
            from_page: Başlangıç satırı (0-based, header sonrası)
            to_page: Bitiş satırı (dahil değil)
        
        Returns:
            List[str] - Her satır için content string
        """
        # Workbook'u yükle
        if binary:
            wb = load_workbook(BytesIO(binary), data_only=False)
        else:
            wb = load_workbook(filename, data_only=False)
        
        all_contents = []
        
        # Her sheet'i işle
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            contents = self._process_sheet(sheet, from_page, to_page)
            all_contents.extend(contents)
        
        return all_contents
    
    def _process_sheet(self, sheet, from_page: int, to_page: int) -> List[str]:
        """Bir sheet'i işler"""
        contents = []
        
        rows = list(sheet.rows)
        if not rows:
            return contents
        
        # İlk iki satırdan EK bilgisini ve liste adını al
        ek_code = self._get_ek_code(rows[0], sheet) if len(rows) > 0 else ""
        list_name = self._get_list_name(rows[1], sheet) if len(rows) > 1 else ""
        
        # Header'ları tespit et (3. satır)
        headers, header_rows = HeaderDetector.detect_headers(sheet)
        
        if not headers:
            return contents
        
        # Data satırlarını işle
        data_row_start = header_rows
        
        for row_idx, row in enumerate(rows[data_row_start:], start=data_row_start):
            # from_page ve to_page kontrolü (header'dan sonraki satır index'i)
            relative_row_idx = row_idx - data_row_start
            
            if relative_row_idx < from_page:
                continue
            if relative_row_idx >= to_page:
                break
            
            # Row'u işle
            content = RowProcessor.process_row(row, headers, sheet, ek_code, list_name, row_idx)
            
            if content:
                contents.append(content)
        
        return contents
    
    def _get_ek_code(self, first_row, sheet) -> str:
        """İlk satırdan EK kodunu al (örn: 'EK-1/C')"""
        if not first_row:
            return ""
        
        for cell in first_row:
            processed_value = CellProcessor.process_cell(cell, sheet)
            if processed_value:
                return processed_value.strip()
        return ""
    
    def _get_list_name(self, second_row, sheet) -> str:
        """İkinci satırdan liste adını al ve küçük harfe çevir"""
        if not second_row:
            return ""

        for cell in second_row:
            processed_value = CellProcessor.process_cell(cell, sheet)
            if processed_value:
                # Türkçe karakterleri düzgün handle et
                text = processed_value.strip()
                # Unicode normalization (combining characters'ı düzelt)
                text = unicodedata.normalize('NFC', text)  # ✅
                # Türkçe-aware lowercase
                text = text.replace('İ', 'i').replace('I', 'ı')  # ✅
                text = text.replace('Ş', 'ş').replace('Ğ', 'ğ')
                text = text.replace('Ü', 'ü').replace('Ö', 'ö').replace('Ç', 'ç')
                return text.lower()
        return ""
        
        """
        for cell in second_row:
            processed_value = CellProcessor.process_cell(cell, sheet)
            if processed_value:
                return processed_value.strip().lower()
        return ""
        """


# ============================================================================
# RAG CHUNK INTERFACE
# ============================================================================

def chunk(filename: str, binary: bytes = None, from_page: int = 0, to_page: int = 100000, 
          lang: str = "Turkish", callback=None, **kwargs) -> List[Dict]:
    """
    SUT XLSX dosyasını RAG için chunk'lara ayırır.
    
    Pattern: sut.py (DOCX) ile aynı output formatı
    
    Args:
        filename: Dosya adı
        binary: Dosya binary (BytesIO için)
        from_page: Başlangıç satırı (0-based, header sonrası)
        to_page: Bitiş satırı (dahil değil)
        lang: Dil (default: "Turkish")
        callback: Progress callback (callback(progress: float, message: str))
        **kwargs: Ekstra parametreler
    
    Returns:
        List[Dict] - Her satır için:
        {
            "docnm_kwd": filename,
            "title_tks": [...],
            "title_sm_tks": [...],
            "content_with_weight": "EK-1/C ...; SIRA NO:1; ...",
            "content_ltks": [...],
            "content_sm_ltks": [...]
        }
    """
    # Progress callback
    if callback:
        callback(0.1, "SUT XLSX parsing started...")
    
    # Document metadata
    doc = {
        "docnm_kwd": filename,
        "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename))
    }
    doc["title_sm_tks"] = rag_tokenizer.fine_grained_tokenize(doc["title_tks"])
    
    # SUT XLSX mi kontrol et
    if not re.search(r"\.xlsx$", filename, re.IGNORECASE):
        raise NotImplementedError("SUT XLSX parser only supports .xlsx files")
    
    # Parse et
    try:
        parser = SutXlsx()
        raw_chunks = parser(filename, binary, from_page, to_page)
        
        if callback:
            callback(0.6, f"Parsed {len(raw_chunks)} rows. Creating chunks...")
        
        # Her chunk için RAG formatına çevir
        result = []
        for i, chunk_content in enumerate(raw_chunks):
            d = copy.deepcopy(doc)
            
            # Content field (sut.py gibi tek field)
            d["content_with_weight"] = chunk_content
            
            
            # Tokenization
            d["content_ltks"] = rag_tokenizer.tokenize(chunk_content)
            d["content_sm_ltks"] = rag_tokenizer.fine_grained_tokenize(d["content_ltks"])
            
            result.append(d)
            
            # Progress update (her 100 chunk'ta bir)
            if callback and i % 100 == 0 and i > 0:
                progress = 0.6 + (i / len(raw_chunks)) * 0.3
                callback(progress, f"Tokenizing chunks... {i}/{len(raw_chunks)}")
        
        if callback:
            callback(0.9, f"Completed. {len(result)} chunks created.")
        
        return result
    
    except Exception as e:
        if callback:
            callback(-1, f"Error parsing SUT XLSX: {str(e)}")
        raise


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    """Standalone test için"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python table_sut.py <xlsx_file>")
        sys.exit(1)
    
    test_file = sys.argv[1]
    
    print("=" * 80)
    print("SUT XLSX Parser Test")
    print("=" * 80)
    print(f"File: {test_file}")
    print()
    
    def progress_callback(p, msg):
        print(f"[{p*100:.1f}%] {msg}")
    
    try:
        chunks = chunk(test_file, callback=progress_callback)
        
        print()
        print("=" * 80)
        print(f"Total chunks: {len(chunks)}")
        print("=" * 80)
        print()
        print("Sample chunks (first 3):")
        print("-" * 80)
        
        for i, chunk in enumerate(chunks[:3], 1):
            content = chunk.get("content_with_weight", "")
            if len(content) > 150:
                content = content[:150] + "..."
            print(f"\n{i}. {content}")
        
        print()
        print("=" * 80)
    
    except Exception as e:
        print(f"\n[ERROR] {str(e)}")
        import traceback
        traceback.print_exc()